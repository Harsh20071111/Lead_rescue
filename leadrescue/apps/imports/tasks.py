import collections
import io
import math
import logging
from contextlib import contextmanager
import requests
import pandas as pd
from celery import shared_task
from django.db import transaction

from apps.imports.models import ImportJob
from apps.leads.models import Activity, Lead
from apps.properties.models import Property
from apps.whatsapp.services.qualification import parse_budget, parse_bhk

logger = logging.getLogger(__name__)


def _stream_xlsx_rows(file_bytes, chunk_size=100):
    """Stream xlsx rows in chunks using openpyxl iter_rows (memory-efficient)."""
    import openpyxl
    wb = openpyxl.load_workbook(file_bytes, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows())]
    chunk = []
    for row in ws.iter_rows(values_only=True):
        chunk.append(dict(zip(headers, row)))
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk)
            chunk = []
    if chunk:
        yield pd.DataFrame(chunk)
    wb.close()


def _stream_xls_rows(file_bytes, chunk_size=100):
    """Stream old xls rows in chunks using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes.read())
    ws = wb.sheet_by_index(0)
    headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
    chunk = []
    for r in range(1, ws.nrows):
        row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
        chunk.append(row)
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk)
            chunk = []
    if chunk:
        yield pd.DataFrame(chunk)

@shared_task(bind=True)
def process_import_job(self, job_id):
    try:
        job = ImportJob.objects.get(id=job_id)
    except ImportJob.DoesNotExist:
        logger.error(f"ImportJob {job_id} not found.")
        return

    if job.status not in (ImportJob.Status.PENDING, ImportJob.Status.MAPPING):
        return

    job.status = ImportJob.Status.PROCESSING
    job.save(update_fields=['status'])

    try:
        with _open_import_file(job) as file_bytes:
            filename = job.file.name.lower()
            if filename.endswith('.csv'):
                reader = pd.read_csv(file_bytes, chunksize=100)
                for chunk in reader:
                    job.refresh_from_db()
                    if job.status == ImportJob.Status.CANCELED:
                        return
                    _process_chunk(job, chunk)
            elif filename.endswith('.xlsx'):
                for chunk in _stream_xlsx_rows(file_bytes):
                    job.refresh_from_db()
                    if job.status == ImportJob.Status.CANCELED:
                        return
                    _process_chunk(job, chunk)
            elif filename.endswith('.xls'):
                for chunk in _stream_xls_rows(file_bytes):
                    job.refresh_from_db()
                    if job.status == ImportJob.Status.CANCELED:
                        return
                    _process_chunk(job, chunk)
            else:
                job.status = ImportJob.Status.FAILED
                job.error_log.append({"error": "Unsupported file format."})
                job.save(update_fields=['status', 'error_log'])
                return
    except Exception as e:
        logger.exception("Failed to read/process file for job %s", job_id)
        job.status = ImportJob.Status.FAILED
        job.error_log.append({"error": f"File read error: {str(e)}"})
        job.save(update_fields=['status', 'error_log'])
        return

    # Determine final status
    job.refresh_from_db()
    if job.status == ImportJob.Status.CANCELED:
        return

    if job.total_rows > 0 and job.successful_rows == 0:
        job.status = ImportJob.Status.FAILED
    else:
        job.status = ImportJob.Status.COMPLETED

    job.save(update_fields=['status'])


@contextmanager
def _open_import_file(job):
    if job.file_url.startswith(("http://", "https://")):
        response = requests.get(job.file_url, timeout=30)
        response.raise_for_status()
        yield io.BytesIO(response.content)
        return

    if not job.file:
        raise ValueError("No import file saved for this job.")

    with job.file.open("rb") as import_file:
        yield import_file


def clean_value(val):
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()


def _process_chunk(job, chunk):
    mapping = job.column_mapping
    chunk = chunk.fillna('')
    
    total = len(chunk)
    successful = 0
    failed = 0
    new_errors = []

    for index, row in chunk.iterrows():
        try:
            if job.target_model == ImportJob.TargetModel.LEAD:
                success, error = _process_lead_row(job, row, mapping)
            elif job.target_model == ImportJob.TargetModel.PROPERTY:
                success, error = _process_property_row(job, row, mapping)
            else:
                success, error = False, "Unknown target model"
            
            if success:
                successful += 1
            else:
                failed += 1
                new_errors.append({"row": index + 2, "error": error})  # +2 because header + 0-index
        except Exception as e:
            logger.exception("Failed to process row %s in job %s", index, job.id)
            failed += 1
            new_errors.append({"row": index + 2, "error": str(e)})
            
    with transaction.atomic():
        job.refresh_from_db()
        job.total_rows += total
        job.processed_rows += total
        job.successful_rows += successful
        job.failed_rows += failed
        job.error_log.extend(new_errors)
        job.save(update_fields=['total_rows', 'processed_rows', 'successful_rows', 'failed_rows', 'error_log'])


def _process_lead_row(job, row, mapping):
    name_col = mapping.get('name')
    phone_col = mapping.get('phone')
    
    if not name_col or not phone_col:
        return False, "Missing mandatory mapping: name or phone"

    name = clean_value(row.get(name_col))
    phone = clean_value(row.get(phone_col))

    if not name or not phone:
        return False, "Name or Phone is empty"

    if Lead.objects.for_agency(job.agency).filter(phone=phone).exists():
        return False, f"Lead with phone {phone} already exists."

    budget_raw = clean_value(row.get(mapping.get('budget', '')))
    bhk_raw = clean_value(row.get(mapping.get('bhk', '')))
    location_raw = clean_value(row.get(mapping.get('location', '')))
    source_raw = clean_value(row.get(mapping.get('source', '')))
    email_raw = clean_value(row.get(mapping.get('email', '')))
    status_raw = clean_value(row.get(mapping.get('status', '')))
    notes_raw = clean_value(row.get(mapping.get('notes', '')))
    
    budget_min, budget_max = parse_budget(budget_raw) if budget_raw else (None, None)
    preferred_bhk = parse_bhk(bhk_raw) if bhk_raw else None

    # Resolve choices
    source_val = Lead.LeadSource.IMPORT
    for choice in Lead.LeadSource.choices:
        if str(choice[1]).lower() == source_raw.lower() or str(choice[0]).lower() == source_raw.lower():
            source_val = choice[0]
            break

    status_val = Lead.LeadStatus.NEW
    for choice in Lead.LeadStatus.choices:
        if str(choice[1]).lower() == status_raw.lower() or str(choice[0]).lower() == status_raw.lower():
            status_val = choice[0]
            break

    lead = Lead.objects.create(
        agency=job.agency,
        assigned_agent=job.initiated_by,
        name=name,
        phone=phone,
        email=email_raw,
        source=source_val,
        status=status_val,
        budget=budget_raw,
        budget_min=budget_min,
        budget_max=budget_max,
        bhk_preference=bhk_raw,
        preferred_bhk=preferred_bhk,
        preferred_location=location_raw,
        area_preference=location_raw,
        notes=notes_raw,
    )
    Activity.objects.create(
        agency=job.agency,
        lead=lead,
        agent=job.initiated_by,
        activity_type=Activity.ActivityType.NOTE,
        content=f"Imported: {name}",
    )
    return True, ""


def _process_property_row(job, row, mapping):
    title_col = mapping.get('title')
    price_col = mapping.get('price')
    city_col = mapping.get('city')
    
    if not title_col or not price_col or not city_col:
        return False, "Missing mandatory mapping: title, price, or city"

    title = clean_value(row.get(title_col))
    price_str = clean_value(row.get(price_col))
    city = clean_value(row.get(city_col))
    locality = clean_value(row.get(mapping.get('locality', '')))

    if not title or not price_str or not city:
        return False, "Title, Price, or City is empty"
    
    try:
        price_str = price_str.replace(',', '').strip()
        price = float(price_str)
    except ValueError:
        return False, f"Invalid price format: {price_str}"

    if Property.objects.for_agency(job.agency).filter(title=title, locality=locality, city=city).exists():
        return False, f"Property '{title}' in {locality}, {city} already exists."

    bhk_raw = clean_value(row.get(mapping.get('bhk', '')))
    bhk = parse_bhk(bhk_raw) if bhk_raw else None
    desc = clean_value(row.get(mapping.get('description', '')))
    
    Property.objects.create(
        agency=job.agency,
        assigned_agent=job.initiated_by,
        title=title,
        price=price,
        city=city,
        locality=locality,
        bhk=bhk,
        description=desc,
        status=Property.PropertyStatus.AVAILABLE,
    )
    return True, ""
